from qtpy.QtWidgets import QMainWindow, QInputDialog
from app.ui.table_page import TablePage
from qtpy.QtWidgets import QTabWidget
from app.model.WI_model import WITableModel
from app.model.input_model import InputTableModel
from app.model.recap_model import RecapTableModel
from qtpy.QtWidgets import QAction
from qtpy.QtCore import Qt
import json
from qtpy.QtWidgets import QFileDialog, QMessageBox


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("WIS")
        self.resize(900, 600)

        # Example pages
        self.models = {
            "Input": InputTableModel(45),
            "Rekap": RecapTableModel(),
        }
        self.models["WI"] = WITableModel()
        self.models["Input"].wi_model = self.models["WI"]
        self.models["WI"].set_input_model(self.models["Input"])


        tabs = QTabWidget()
        self.setCentralWidget(tabs)
        for name, model in self.models.items():
            if name == "Input":
                tabs.addTab(
                    TablePage(
                        model,
                        wi_model=self.models["WI"],
                    ),
                    name
                )
            elif name == "Rekap":
                tabs.addTab(
                    TablePage(
                        model,
                        is_readonly=True
                    ),
                    name
                )
            else:
                tabs.addTab(TablePage(model), name)

        self.current_file = None
        self.is_dirty = False
        self.jp_duration = 45
        self._create_file_menu()
        self._create_settings_menu()
        self._connect_dirty_signals()
        self.new_file()

        self.models["Rekap"].get_wi_name(self.models["WI"]._data)

        self.models["Input"].validationFailed.connect(
                self.show_validation_error
            )

        self.models["WI"].nameChanged.connect(
                self.models["Input"].update_wi_name
            )

        self.models["WI"].WIChanged.connect(
                self.update_rekap
            )

        self.models["Input"].inputChanged.connect(
                self.update_rekap
            )

    def update_rekap(self):
        self.models["Rekap"].get_wi_name(self.models["WI"]._data)
        self.models["Rekap"].get_recap(
                self.models["Input"]._data, 
                self.models["WI"]._data)

    def _create_file_menu(self):
        file_menu = self.menuBar().addMenu("File")

        new_act = QAction("New", self)
        new_act.setShortcut("Ctrl+N")
        new_act.triggered.connect(self.new_file)

        open_act = QAction("Open…", self)
        open_act.setShortcut("Ctrl+O")
        open_act.triggered.connect(self.open_file)

        save_act = QAction("Save", self)
        save_act.setShortcut("Ctrl+S")
        save_act.triggered.connect(self.save_file)

        save_as_act = QAction("Save As…", self)
        save_as_act.setShortcut("Ctrl+Shift+S")
        save_as_act.triggered.connect(self.save_file_as)

        file_menu.addSeparator()

        export_excel_act = QAction("Export Excel…", self)
        export_excel_act.triggered.connect(self.export_to_excel)

        file_menu.addActions([
            new_act,
            open_act,
            save_act,
            save_as_act,
            export_excel_act,
        ])

    def _create_settings_menu(self):
        settings_menu = self.menuBar().addMenu("Settings")

        edit_jp_act = QAction("Edit JP duration...", self)
        edit_jp_act.triggered.connect(self.edit_jp_duration)

        settings_menu.addAction(edit_jp_act)

    def edit_jp_duration(self):
        new_duration, ok = QInputDialog.getInt(
            self,
            "Edit JP",
            "Panjang JP (Menit):",
            value=self.jp_duration,
            minValue=1,
        )

        if ok:
            self.jp_duration = new_duration
            self.models["Input"].set_jp_duration(new_duration)
            self._mark_dirty()

    def new_file(self):
        if not self._maybe_save():
            return

        for model in self.models.values():
            model.clear()

        self.current_file = None
        self.is_dirty = False
        self.jp_duration = 45
        self._update_title()

    def save_file(self):
        if self.current_file is None:
            self.save_file_as()
            if self.current_file is None:  # User cancelled Save As
                return

        data = {
            "jp_duration": self.jp_duration,
            "models": {
                name: model.to_json()
                for name, model in self.models.items()
            }
        }

        with open(self.current_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        self.is_dirty = False
        self._update_title()

    def save_file_as(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save As",
            "",
            "WIS Files (*.json)"
        )

        if not path:
            return

        if not path.endswith(".json"):
            path += ".json"

        self.current_file = path
        self.save_file()

    def open_file(self):
        if not self._maybe_save():
            return
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open File",
            "",
            "WIS Files (*.json)"
        )

        if not path:
            return

        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.jp_duration = data.get("jp_duration", 45)
            self.models["Input"].set_jp_duration(self.jp_duration)

            model_data = data.get("models", data)

            for name, model in self.models.items():
                if name in model_data:
                    model.from_json(model_data[name])

            self.current_file = path
            self.is_dirty = False
            self._update_title()

        except Exception as e:
            QMessageBox.critical(self, "Open failed", str(e))

    def _update_title(self):
        name = self.current_file or "Untitled"
        star = "*" if self.is_dirty else ""
        self.setWindowTitle(f"WIS — {name}{star}")

    def show_validation_error(self, message):
        QMessageBox.warning(self, "Validation Error", message)

    def _connect_dirty_signals(self):
        for name, model in self.models.items():
            model.dataChanged.connect(self._mark_dirty)
            model.rowsInserted.connect(self._mark_dirty)
            model.rowsRemoved.connect(self._mark_dirty)
            model.modelReset.connect(self._mark_dirty)

    def _mark_dirty(self, *args, **kwargs):
        if not self.is_dirty:
            self.is_dirty = True
            self._update_title()

    def closeEvent(self, event):
        if self._maybe_save():
            event.accept()
        else:
            event.ignore()

    def _maybe_save(self) -> bool:
        if not self.is_dirty:
            return True

        res = QMessageBox.question(
            self,
            "Unsaved changes",
            "You have unsaved changes. Save them?",
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel
        )

        if res == QMessageBox.Save:
            self.save_file()
            return not self.is_dirty  # in case user canceled Save As

        if res == QMessageBox.Discard:
            return True

        return False

    def export_to_excel(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "Error", "openpyxl is not installed. Please install it using: pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Excel",
            "",
            "Excel Files (*.xlsx)"
        )

        if not path:
            return

        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = openpyxl.Workbook()
            # Remove the default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            for name, model in self.models.items():
                sheet = wb.create_sheet(title=name)

                # Write headers
                headers = [model.headerData(i, Qt.Horizontal) for i in range(model.columnCount())]
                sheet.append(headers)

                # Write data
                for r in range(model.rowCount()):
                    row_data = [model.data(model.index(r, c)) for c in range(model.columnCount())]
                    sheet.append(row_data)
            
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Data successfully exported to {path}")

        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
